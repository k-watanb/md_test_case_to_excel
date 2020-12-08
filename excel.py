# coding: utf-8

import sys
from itertools import product

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.borders import BORDER_THIN


def write_summary(df: pd.DataFrame, writer: pd.ExcelWriter, sheet_name: str = "Summary") -> None:
    df_summary = pd.DataFrame(index=["正常", "異常"], columns=["OK", "NG", "--"])
    for row, col in product(list(df_summary.index), list(df_summary.columns)):
        df_summary[col][row] = ((df["pos-neg"] == row) & (df["result"] == col)).sum()

    df_summary.to_excel(writer, sheet_name=sheet_name)
    worksheet = writer.sheets[sheet_name]
    worksheet.insert_rows(0)
    worksheet.insert_cols(0)


def write_test_specification(df: pd.DataFrame, writer: pd.ExcelWriter, config_excel: dict, merge_cells: bool,
                             sheet_name: str = "TestSpecification") -> None:
    """
    Notes:
        エクセル列と項目名は以下の様に対応しています。
        A: item, B: sub-item, C: pos-neg, D: sub-sub-item, E: result, F: No, G: steps, H: expected, I: notes

    """
    try:
        col_names = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[0: len(config_excel["name"])]
    except IndexError:
        sys.stderr.write("[ERROR] 列は26列以下にしてください")
        sys.exit(1)

    df_excel = df.copy()
    df_excel.rename(columns=config_excel["name"], inplace=True)

    # マージできるようにマルチインデックス化
    df_excel.set_index([str(v) for k, v in config_excel["name"].items() if k in config_excel["index"]], inplace=True)

    # 項目番号をカラムとして追加しておく
    df_excel.to_excel(writer, sheet_name=sheet_name, merge_cells=merge_cells)

    # ここからExcelデータの見た目を整えていく
    worksheet = writer.sheets[sheet_name]

    # ヘッダーのスタイル
    for col in col_names:
        __col_address = col + "1"
        worksheet[__col_address].font = Font(name=config_excel["font"], b=True, color="ffffff")
        worksheet[__col_address].alignment = Alignment(
            vertical="center",
            horizontal="center",
            wrap_text=True
        )
        worksheet[__col_address].fill = PatternFill(
            patternType="solid",
            fgColor="4f81bd"
        )

    # 列幅調整
    for col, length in zip(col_names, [v for _, v in config_excel["length"].items()]):
        worksheet.column_dimensions[col].width = length

    # データセルのスタイル調整
    for i in range(len(df_excel)):

        for col, horizontal, vertical in zip(col_names, [v for _, v in config_excel["horizontal"].items()],
                                             [v for _, v in config_excel["vertical"].items()]):
            __col_address = col + f"{i + 2}"
            worksheet[__col_address].alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True)
            worksheet[__col_address].font = Font(name=config_excel["font"])
            worksheet[__col_address].border = Border(left=Side(style=BORDER_THIN),
                                                     right=Side(style=BORDER_THIN),
                                                     top=Side(style=BORDER_THIN),
                                                     bottom=Side(style=BORDER_THIN))


def convert_df_to_excel(df: pd.DataFrame, config_excel: dict, output_path: str = "sample.xlsx",
                        merge_cells: bool = True) -> None:
    """
    convert_md_to_df()により生成されたデータフレームをエクセルファイルに変換します。

    Args:
        df:             convert_md_to_df()により生成されたデータフレーム

    Returns:
        None
    """
    writer = pd.ExcelWriter(output_path)

    write_summary(df, writer)
    write_test_specification(df, writer, config_excel, merge_cells)

    try:
        writer.save()
    except PermissionError:
        sys.stderr.write("[ERROR] 仕様書のエクセルファイルを閉じてください")
        sys.exit(1)
