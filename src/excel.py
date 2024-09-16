from itertools import product
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from src.config_loader import Config, load_column_names


def apply_cell_style(cell, font, fill=None, alignment=None, border=None):
    cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


class ExcelWriter:

    def __init__(self, df: pd.DataFrame, config_excel: Config):
        self.df = df.copy()
        self.config = config_excel

        self.columns = load_column_names(self.config)
        self.col_names = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[:len(self.columns)]
        if len(self.col_names) < len(self.columns):
            raise IndexError("列は26列以下にしてください")

    def __write_summary_sheet(self, writer: pd.ExcelWriter):
        """クロス集計表をエクセルシートに書き込みます。

        Args:
            writer (pd.ExcelWriter):    ExcelWriterオブジェクト
            sheet_name (str):           シート名
        """

        # クロス集計表を作成
        df_summary = pd.crosstab(self.df["pos_neg"], self.df["result"])
        df_summary = df_summary.rename_axis("").reset_index()

        # エクセルに書き込む
        df_summary.to_excel(writer, sheet_name=self.config.excel_settings.sheet_name.summary, index=False)

        # ワークシートの取得と行・列の挿入
        worksheet = writer.sheets[self.config.excel_settings.sheet_name.summary]
        worksheet.insert_rows(0)
        worksheet.insert_cols(0)

        # レイアウト調整（ヘッダー色）
        header_font = Font(name=self.config.excel_settings.font_name, bold=True, color="ffffff")
        header_fill = PatternFill(patternType="solid", fgColor="4f81bd")
        header_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        border_style = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # ヘッダー（列）
        for i in range(1, len(df_summary.columns) + 1):
            apply_cell_style(
                worksheet[f"{self.col_names[i]}2"],
                font=header_font,
                fill=header_fill,
                alignment=header_alignment,
                border=border_style,
            )

        # ヘッダー（行）
        for i in range(2, len(df_summary.index) + 3):
            apply_cell_style(
                worksheet[f"B{i}"],
                font=header_font,
                fill=header_fill,
                alignment=header_alignment,
                border=border_style,
            )

        # データセルのスタイル調整
        for i, j in product(range(3, len(df_summary.index) + 3), range(2, len(df_summary.columns) + 1)):
            apply_cell_style(
                worksheet[f"{self.col_names[j]}{i}"],
                font=Font(name=self.config.excel_settings.font_name),
                border=border_style,
            )

    def __write_test_specification_sheet(self,
                                         writer: pd.ExcelWriter,
                                         merge_cells: bool = False,
                                         ):
        """テスト仕様書をエクセルシートに書き込みます。

        Args:
            writer (pd.ExcelWriter):    ExcelWriterオブジェクト
            merge_cells (bool):         セルをマージするかどうか
        """
        df_excel = self.df.copy()
        df_excel.columns = self.columns

        # マージできるようにマルチインデックス化
        if merge_cells:
            merge_columns = [v["name"] for k, v in self.config.columns.model_dump().items() if v["multi_idx"]]
            df_excel = df_excel.set_index([v for v in list(df_excel.columns) if v in merge_columns])

        # 項目番号をカラムとして追加しておく
        df_excel.to_excel(
            writer, sheet_name=self.config.excel_settings.sheet_name.test, merge_cells=merge_cells, index=merge_cells,
        )

        # ワークシートの取得
        worksheet = writer.sheets[self.config.excel_settings.sheet_name.test]

        # ヘッダーのスタイル設定
        for col, length in zip(self.col_names, [col["length"] for col in self.config.columns.model_dump().values()]):
            apply_cell_style(
                worksheet[f"{col}1"],
                font=Font(name=self.config.excel_settings.font_name, bold=True, color="ffffff"),
                fill=PatternFill(patternType="solid", fgColor="4f81bd"),
                alignment=Alignment(vertical="center", horizontal="center", wrap_text=True),
            )

            # 列幅調整
            worksheet.column_dimensions[col].width = length

        # データセルのスタイル調整
        for i in range(len(df_excel)):
            for col, (horizontal, vertical) in (
                    zip(self.col_names, zip([x["horizontal"] for x in self.config.columns.model_dump().values()],
                                            [x["vertical"] for x in self.config.columns.model_dump().values()])
                        )):
                apply_cell_style(
                    worksheet[f"{col}{i + 2}"],
                    font=Font(name=self.config.excel_settings.font_name),
                    alignment=Alignment(horizontal=horizontal, vertical=vertical, wrap_text=True),
                    border=Border(left=Side(style="thin"), right=Side(style="thin"),
                                  top=Side(style="thin"), bottom=Side(style="thin"))
                )

    def __call__(self, output_path: Path, merge_cells: bool = True):
        """
        convert_md_to_df()により生成されたデータフレームをエクセルファイルに変換します。

        Args:
            df (pd.DataFrame):    convert_md_to_df()により生成されたデータフレーム
        """
        try:
            with pd.ExcelWriter(output_path) as writer:
                self.__write_summary_sheet(writer)
                self.__write_test_specification_sheet(writer, merge_cells)
        except PermissionError:
            raise PermissionError(f"出力先のファイルを開いている可能性があります。エクセルファイルを閉じてください。")
        except Exception as e:
            raise ValueError(f"エクセルファイル出力中に不明なエラーが発生しました：\n{e}")

        return output_path
